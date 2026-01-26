
import openpyxl
import os
import json

excel_path = r"c:\Users\Victus\ndata\src\excel\excel.xlsx"
output_path = r"c:\Users\Victus\ndata\excel_content.json"

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    content = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        rows = []
        for row in sheet.iter_rows(max_row=50, max_col=20, values_only=True):
            if any(cell is not None for cell in row):
                rows.append(row)
        content[sheetname] = rows
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(content, f, indent=2, default=str)
    print(f"Content exported to {output_path}")
else:
    print("Excel file not found")
