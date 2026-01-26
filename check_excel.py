
try:
    import openpyxl
    print("openpyxl is installed")
except ImportError:
    print("openpyxl is not installed")

import os

excel_path = r"c:\Users\Victus\ndata\src\excel\excel.xlsx"
if os.path.exists(excel_path):
    print(f"Excel file found at {excel_path}")
    if 'openpyxl' in globals() or 'openpyxl' in locals():
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        for sheetname in wb.sheetnames:
            print(f"\nSheet: {sheetname}")
            sheet = wb[sheetname]
            for row in sheet.iter_rows(max_row=10, values_only=True):
                print(row)
else:
    print("Excel file not found")
